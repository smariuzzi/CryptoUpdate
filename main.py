from openpyxl import load_workbook

file_name = 'excel.xlsx'
sheet_name = 'Sheet1'
cell = 'A1'


def read_cell(file_name, sheet_name, cell):
    wb = load_workbook(filename=file_name)
    ws = wb[sheet_name]
    return ws[cell].value


def write_cell(file_name, sheet_name, cell, info):
    wb = load_workbook(filename=file_name)
    ws = wb[sheet_name]
    ws[cell] = info
    wb.save(file_name)


def read_key(file_name='keys.txt'):
    try:
        with open(file_name, mode='r') as file:
            key = file.read()
    except FileNotFoundError:
        text = 'Inserire la chiave per CoinMarketCap: '
        key = input(text)
        with open(file_name, mode='w') as file:
            file.write(key)
    return key


def get_column(index):
    diff = ord('Z') - ord('A') + 1
    times = (index-1) // diff
    index = index - times * diff
    letter = lambda x: chr(x + ord('A') - 1)
    return letter(times) + letter(index) if times else letter(index)


def find_cells(workbook, keys):
    wb = workbook
    sheets_names = wb.sheetnames
    cells = {}
    for sheet_name in sheets_names:
        sheet = wb[sheet_name]
        for row in range(1, sheet.max_row + 1):
            for column in range(1, sheet.max_column + 1):
                cell = '{}{}'.format(get_column(column), row)
                value = sheet[cell].value
                if value in keys:
                    cells[value] = [sheet_name, cell]
                    keys.remove(value)
    return cells


if __name__ == '__main__':
    wb = load_workbook(filename=file_name)
    keys = ['BTC', 'ETH', 'STE', 'BHO']
    print(find_cells(wb, keys))


